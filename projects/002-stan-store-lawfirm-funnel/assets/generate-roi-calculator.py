"""Generate ROI Calculator Excel file for Automated Law Firm Blueprint.

v2: All percentages entered as whole numbers (91 = 91%). Formulas divide by 100.
    Cell comments (tooltips) on every input. Data validation on inputs.
"""
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# Colors
DARK_BG = "1A1A1A"
LIGHT_TEXT = "F0F0F0"
INPUT_BG = "FFF9E6"
SUMMARY_BG = "E6F7E6"
SECTION_BG = "2A2A2A"

header_font = Font(name="Inter", size=14, bold=True, color=LIGHT_TEXT)
section_font = Font(name="Inter", size=11, bold=True, color=LIGHT_TEXT)
label_font = Font(name="Inter", size=10, bold=True)
data_font = Font(name="Inter", size=10)
note_font = Font(name="Inter", size=9, italic=True, color="666666")
input_fill = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
dark_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
section_fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
summary_fill = PatternFill(start_color=SUMMARY_BG, end_color=SUMMARY_BG, fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

AUTHOR = "ROI Calculator"


def style_header(ws, row, cols=3):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = dark_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def style_section(ws, row, cols=3):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = section_fill
        cell.font = section_font


def add_tooltip(ws, row, col, text):
    """Add a hover tooltip (comment) to a cell."""
    ws.cell(row=row, column=col).comment = Comment(text, AUTHOR, width=300, height=120)


def add_validation(ws, row, col, min_val, max_val, msg):
    """Add data validation to a single cell."""
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(min_val),
        formula2=str(max_val),
        allow_blank=True,
    )
    dv.error = f"Please enter a number between {min_val} and {max_val}."
    dv.errorTitle = "Invalid Input"
    dv.prompt = msg
    dv.promptTitle = "Input Required"
    dv.showInputMessage = True
    dv.showErrorMessage = True
    cell_ref = f"B{row}"
    dv.add(cell_ref)
    ws.add_data_validation(dv)


# ============================================================
# TAB 1: Your Firm Profile
# ============================================================
ws1 = wb.active
ws1.title = "Your Firm Profile"
ws1.column_dimensions["A"].width = 42
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 32

# Header
ws1["A1"] = "Your Firm Profile"
style_header(ws1, 1)
ws1.row_dimensions[1].height = 30

# Column headers
for col, val in [(1, "Field Name"), (2, "Your Value"), (3, "Units / Notes")]:
    cell = ws1.cell(row=2, column=col, value=val)
    cell.font = Font(name="Inter", size=10, bold=True, color=LIGHT_TEXT)
    cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

# Field definitions: (name, units_text, tooltip, min, max, is_section)
fields = [
    ("BILLING & REVENUE", None, None, None, None, True),
    (
        "Average Billing Rate",
        "$/hour",
        "Your standard hourly billing rate in dollars.\nExample: If you bill $275/hour, enter 275.\nThis is used to calculate the dollar value of recovered time.",
        50, 1000, False,
    ),
    (
        "Number of Attorneys",
        "Count",
        "Total number of attorneys at your firm (partners + associates).\nParalegals are NOT included.\nUsed to multiply per-attorney time savings across the firm.",
        1, 100, False,
    ),
    (
        "Average Revenue Per Case",
        "$",
        "Average total revenue you collect per case from start to close.\nInclude all fees (retainer + hourly + flat fee).\nExample: If a typical bankruptcy case brings in $3,000, enter 3000.",
        500, 100000, False,
    ),
    (
        "% of Billable Time Actually Billed",
        "Enter as whole number (e.g. 91 for 91%)",
        "What percentage of your attorneys' billable work actually gets invoiced?\nIndustry average is 70-85%. Top firms hit 90-95%.\nEnter as a whole number: 91 means 91%.\nThe 'missing' percentage is billing leakage — work done but never billed.",
        50, 100, False,
    ),
    ("CASELOAD & WORKLOAD", None, None, None, None, True),
    (
        "Current Caseload (active cases)",
        "Count",
        "Total number of currently active/open cases across the firm.\nInclude all matters in progress, not just new cases.\nUsed to estimate billing capture improvements.",
        5, 500, False,
    ),
    (
        "Hours Lost to Admin Per Attorney/Week",
        "Hours (2-10)",
        "How many hours per week does each attorney spend on NON-billable admin?\nIncludes: filing, data entry, scheduling, document formatting, status updates.\nDoes NOT include: client meetings, legal research, court time.\nIndustry average: 6-8 hours/week. Be honest — most firms underestimate this.",
        1, 20, False,
    ),
    ("MARKETING & INTAKE", None, None, None, None, True),
    (
        "New Leads Per Month",
        "Count",
        "Total new potential client inquiries per month across ALL channels.\nIncludes: phone calls, web forms, emails, referrals, walk-ins.\nCount every inquiry, even ones that don't convert.",
        1, 500, False,
    ),
    (
        "Lead Response Time (current)",
        "Hours (1-48)",
        "On average, how many hours between a new lead contacting you and your first response?\nBe honest. Count business hours only.\n1 = same hour. 24 = next business day. 48 = two days.\nFaster response = higher conversion. Under 1 hour is the gold standard.",
        0.5, 72, False,
    ),
    (
        "Lead Conversion Rate (current)",
        "Enter as whole number (e.g. 7 for 7%)",
        "What percentage of leads become paying clients?\nExample: If 50 leads/month and 4 become clients, enter 8 (for 8%).\nEnter as a whole number: 7 means 7%.\nIndustry averages: Bankruptcy ~13%, Criminal Defense ~4.5%, Overall legal ~7%.",
        1, 80, False,
    ),
    (
        "Monthly Marketing Spend",
        "$",
        "Total monthly marketing budget across all channels.\nIncludes: Google Ads, LSAs, social media ads, SEO services, directory listings, etc.\nDoes NOT include: staff time or referral fees.\nEnter 0 if you rely entirely on referrals.",
        0, 500000, False,
    ),
    ("OPERATIONS", None, None, None, None, True),
    (
        'Weekly "Where\'s My Case?" Calls',
        "Count (5-30)",
        "How many times per week do clients call/email asking for case status updates?\nCount across all staff who field these calls.\nThese calls are pure overhead — they don't advance the case.\nAutomated status updates eliminate 80% of these.",
        0, 100, False,
    ),
    (
        "Time Spent on Billing/Collections Per Week",
        "Hours (2-15)",
        "Total hours per week spent on billing tasks across all staff.\nIncludes: time entry, invoice preparation, collections calls, payment processing.\nThis is overhead that can be largely automated.",
        0, 40, False,
    ),
    (
        "Missed Deadlines in Past 12 Months",
        "Count (0-5+)",
        "Number of court deadlines, filing deadlines, or statute of limitations missed in the past year.\nBe honest — even 1 missed deadline is a malpractice risk.\n0 is great but rare. 1-2 is common. 3+ means your system is failing.",
        0, 20, False,
    ),
]

row = 3
input_rows = {}
for name, units, tooltip, min_v, max_v, is_section in fields:
    row += 1
    if is_section:
        ws1.cell(row=row, column=1, value=name).font = Font(name="Inter", size=10, bold=True, color="00BFFF")
        style_section(ws1, row)
        continue
    ws1.cell(row=row, column=1, value=name).font = label_font
    ws1.cell(row=row, column=3, value=units).font = note_font

    # Style input cell
    cell = ws1.cell(row=row, column=2)
    cell.fill = input_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

    # Add tooltip to the input cell AND the label cell
    if tooltip:
        add_tooltip(ws1, row, 1, tooltip)
        add_tooltip(ws1, row, 2, tooltip)

    # Add data validation
    if min_v is not None and max_v is not None:
        add_validation(ws1, row, 2, min_v, max_v, tooltip.split("\n")[0] if tooltip else name)

    input_rows[name] = row

# Instructions
row += 2
ws1.cell(row=row, column=1, value="Instructions: Fill in Column B with your firm's current metrics. Hover over any field for a detailed explanation. Percentages should be entered as whole numbers (e.g. 91 for 91%, 20 for 20%). Use practice-area benchmarks in the Summary tab as a starting point.").font = note_font
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

# Map named references — percentages get /100 in formulas
r = input_rows
# Raw cell references (no /100 applied yet — formulas apply /100 where needed)
RAW = {
    "BillingRate": f"'Your Firm Profile'!B{r['Average Billing Rate']}",
    "NumAttorneys": f"'Your Firm Profile'!B{r['Number of Attorneys']}",
    "AvgRevPerCase": f"'Your Firm Profile'!B{r['Average Revenue Per Case']}",
    "BillableCapture_raw": f"'Your Firm Profile'!B{r['% of Billable Time Actually Billed']}",
    "Caseload": f"'Your Firm Profile'!B{r['Current Caseload (active cases)']}",
    "AdminHours": f"'Your Firm Profile'!B{r['Hours Lost to Admin Per Attorney/Week']}",
    "Leads": f"'Your Firm Profile'!B{r['New Leads Per Month']}",
    "ResponseTime": f"'Your Firm Profile'!B{r['Lead Response Time (current)']}",
    "ConvRate_raw": f"'Your Firm Profile'!B{r['Lead Conversion Rate (current)']}",
    "MktSpend": f"'Your Firm Profile'!B{r['Monthly Marketing Spend']}",
    "SupportCalls": "'Your Firm Profile'!B" + str(r["Weekly \"Where's My Case?\" Calls"]),
    "BillingTime": f"'Your Firm Profile'!B{r['Time Spent on Billing/Collections Per Week']}",
    "MissedDeadlines": f"'Your Firm Profile'!B{r['Missed Deadlines in Past 12 Months']}",
}

# Convenience: percentage refs that include /100
PCT_CONV = f"({RAW['ConvRate_raw']}/100)"
PCT_BILL = f"({RAW['BillableCapture_raw']}/100)"

# ============================================================
# TAB 2: Phase-by-Phase ROI
# ============================================================
ws2 = wb.create_sheet("Phase-by-Phase ROI")
ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 20

# Title
ws2["A1"] = "ROI Calculator: Automated Law Firm Blueprint"
style_header(ws2, 1, 5)
ws2.merge_cells("A1:E1")
ws2["A2"] = 'Based on inputs from "Your Firm Profile" tab'
ws2["A2"].font = note_font
ws2["A3"] = "All projections are estimates based on industry benchmarks and may vary."
ws2["A3"].font = note_font


def phase_headers(ws, row):
    headers = ["Metric", "Current", "Improved", "Change", "Annual Impact"]
    tooltips = [
        "What's being measured",
        "Your current state based on inputs",
        "Projected state after automation",
        "The difference (improvement)",
        "Dollar impact over 12 months",
    ]
    for i, (h, tip) in enumerate(zip(headers, tooltips), 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(name="Inter", size=10, bold=True, color=LIGHT_TEXT)
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        cell.comment = Comment(tip, AUTHOR, width=200, height=60)


# --- PHASE 1: Intake + Deadlines ---
R = 5
ws2.cell(row=R, column=1, value="PHASE 1: Intake + Deadlines (Month 1)")
style_section(ws2, R, 5)
ws2.merge_cells(f"A{R}:E{R}")
add_tooltip(ws2, R, 1, "Phase 1 focuses on automating client intake and deadline tracking.\nBiggest ROI driver: faster lead response = higher conversion.\nSecondary benefit: zero missed deadlines = eliminated malpractice risk.")

R += 1
phase_headers(ws2, R)

R += 1  # Current conversion rate
ws2.cell(row=R, column=1, value="Lead Conversion Rate").font = label_font
add_tooltip(ws2, R, 1, "Percentage of leads that become paying clients.\nAutomation improves this by 30% (capped at 60%) through faster response times.\nFaster response = the lead hasn't called your competitor yet.")
ws2.cell(row=R, column=2).value = f"={PCT_CONV}"
ws2.cell(row=R, column=2).number_format = '0%'
ws2.cell(row=R, column=3).value = f"=MIN({PCT_CONV}*1.3,0.6)"
ws2.cell(row=R, column=3).number_format = '0%'
p1_conv_cur = R

R += 1  # Conversions/month
ws2.cell(row=R, column=1, value="Conversions Per Month").font = label_font
add_tooltip(ws2, R, 1, "How many leads turn into clients each month.\nCurrent = Leads × Conversion Rate.\nImproved = Leads × Improved Conversion Rate.")
ws2.cell(row=R, column=2).value = f"={RAW['Leads']}*B{p1_conv_cur}"
ws2.cell(row=R, column=2).number_format = '0.0'
ws2.cell(row=R, column=3).value = f"={RAW['Leads']}*C{p1_conv_cur}"
ws2.cell(row=R, column=3).number_format = '0.0'
p1_conv_mo = R

R += 1  # New cases gained
ws2.cell(row=R, column=1, value="New Cases Gained/Month").font = label_font
add_tooltip(ws2, R, 1, "Additional cases per month from improved conversion.\nThis is the difference between improved and current conversions.")
ws2.cell(row=R, column=4).value = f"=C{p1_conv_mo}-B{p1_conv_mo}"
ws2.cell(row=R, column=4).number_format = '0.0'
p1_new_cases = R

R += 1  # Revenue from new cases
ws2.cell(row=R, column=1, value="Revenue from New Cases/Month").font = label_font
add_tooltip(ws2, R, 1, "Dollar value of additional cases gained.\nNew Cases × Average Revenue Per Case.\nAnnual = Monthly × 12.")
ws2.cell(row=R, column=4).value = f"=D{p1_new_cases}*{RAW['AvgRevPerCase']}"
ws2.cell(row=R, column=4).number_format = '$#,##0'
ws2.cell(row=R, column=5).value = f"=D{R}*12"
ws2.cell(row=R, column=5).number_format = '$#,##0'
p1_rev = R

R += 1  # Missed deadlines
ws2.cell(row=R, column=1, value="Missed Deadlines → Zero").font = label_font
add_tooltip(ws2, R, 1, "Automated deadline tracking eliminates missed filings.\nThe dollar value of avoided malpractice claims is not calculated here\nbut is often the single highest-value benefit of automation.")
ws2.cell(row=R, column=2).value = f"={RAW['MissedDeadlines']}"
ws2.cell(row=R, column=3, value=0)

R += 1  # Phase 1 totals
ws2.cell(row=R, column=1, value="PHASE 1 ANNUAL ROI").font = Font(name="Inter", size=11, bold=True)
ws2.cell(row=R, column=5).value = f"=E{p1_rev}"
ws2.cell(row=R, column=5).number_format = '$#,##0'
ws2.cell(row=R, column=5).fill = summary_fill
ws2.cell(row=R, column=5).font = Font(name="Inter", size=11, bold=True)
p1_annual = R

R += 1
ws2.cell(row=R, column=1, value="Assumption: 30% conversion improvement (capped at 60%). Faster response captures leads before they call competitors.").font = note_font
ws2.merge_cells(f"A{R}:E{R}")

# --- PHASE 2: Document Assembly + Billing ---
R += 2
ws2.cell(row=R, column=1, value="PHASE 2: Document Assembly + Billing (Month 2)")
style_section(ws2, R, 5)
ws2.merge_cells(f"A{R}:E{R}")
add_tooltip(ws2, R, 1, "Phase 2 automates document generation and billing capture.\n65% admin time reduction from template automation.\n15 percentage point billing capture improvement from automated time tracking.")

R += 1
phase_headers(ws2, R)

R += 1  # Admin hours
ws2.cell(row=R, column=1, value="Admin Hours Lost/Attorney/Week").font = label_font
add_tooltip(ws2, R, 1, "Hours each attorney loses to admin tasks weekly.\nAutomation reduces this by 65% (multiplied by 0.35 = 35% of original).\nSaved hours become available for billable work.")
ws2.cell(row=R, column=2).value = f"={RAW['AdminHours']}"
ws2.cell(row=R, column=2).number_format = '0.0'
ws2.cell(row=R, column=3).value = f"={RAW['AdminHours']}*0.35"
ws2.cell(row=R, column=3).number_format = '0.0'
p2_admin = R

R += 1  # Hours recovered
ws2.cell(row=R, column=1, value="Hours Recovered/Week (All Attorneys)").font = label_font
add_tooltip(ws2, R, 1, "Total recovered hours = (Current - Improved) × Number of Attorneys.\nThese hours are now available for billable work.\nAnnual = Weekly × 52 weeks.")
ws2.cell(row=R, column=4).value = f"=(B{p2_admin}-C{p2_admin})*{RAW['NumAttorneys']}"
ws2.cell(row=R, column=4).number_format = '0.0'
ws2.cell(row=R, column=5).value = f"=D{R}*52"
ws2.cell(row=R, column=5).number_format = '0.0'
p2_hours = R

R += 1  # Revenue from recovered hours
ws2.cell(row=R, column=1, value="Revenue from Recovered Hours").font = label_font
add_tooltip(ws2, R, 1, "Dollar value of recovered time = Hours × Billing Rate.\nAssumes recovered hours are used for billable work.\nAnnual = Weekly revenue × 52 weeks.")
ws2.cell(row=R, column=4).value = f"=D{p2_hours}*{RAW['BillingRate']}"
ws2.cell(row=R, column=4).number_format = '$#,##0'
ws2.cell(row=R, column=5).value = f"=D{R}*52"
ws2.cell(row=R, column=5).number_format = '$#,##0'
p2_rev_hours = R

R += 1  # Billing capture (FIXED: /100 for percentage)
ws2.cell(row=R, column=1, value="Billable Capture Rate").font = label_font
add_tooltip(ws2, R, 1, "Percentage of billable work that actually gets invoiced.\nAutomated time tracking improves this by 15 percentage points (capped at 95%).\nExample: 70% current + 15% improvement = 85% improved.")
ws2.cell(row=R, column=2).value = f"={PCT_BILL}"
ws2.cell(row=R, column=2).number_format = '0%'
ws2.cell(row=R, column=3).value = f"=MIN({PCT_BILL}+0.15,0.95)"
ws2.cell(row=R, column=3).number_format = '0%'
p2_capture = R

R += 1  # Additional billed revenue
ws2.cell(row=R, column=1, value="Additional Billed Revenue/Month").font = label_font
add_tooltip(ws2, R, 1, "Revenue recovered from improved billing capture.\nFormula: (Improved% - Current%) × Monthly Cases × Avg Revenue Per Case.\nThis is money you're already earning but currently not billing for.")
ws2.cell(row=R, column=4).value = f"=(C{p2_capture}-B{p2_capture})*({RAW['Caseload']}/12)*{RAW['AvgRevPerCase']}"
ws2.cell(row=R, column=4).number_format = '$#,##0'
ws2.cell(row=R, column=5).value = f"=D{R}*12"
ws2.cell(row=R, column=5).number_format = '$#,##0'
p2_rev_billing = R

R += 1  # Phase 2 totals
ws2.cell(row=R, column=1, value="PHASE 2 ANNUAL ROI").font = Font(name="Inter", size=11, bold=True)
ws2.cell(row=R, column=5).value = f"=E{p2_rev_hours}+E{p2_rev_billing}"
ws2.cell(row=R, column=5).number_format = '$#,##0'
ws2.cell(row=R, column=5).fill = summary_fill
ws2.cell(row=R, column=5).font = Font(name="Inter", size=11, bold=True)
p2_annual = R

R += 1
ws2.cell(row=R, column=1, value="Assumptions: 65% admin reduction, 15 percentage point billing capture improvement (capped at 95%).").font = note_font
ws2.merge_cells(f"A{R}:E{R}")

# --- PHASE 3: Client Communication ---
R += 2
ws2.cell(row=R, column=1, value="PHASE 3: Client Communication (Month 3)")
style_section(ws2, R, 5)
ws2.merge_cells(f"A{R}:E{R}")
add_tooltip(ws2, R, 1, "Phase 3 implements client portals and automated status updates.\n80% reduction in 'where is my case?' calls.\nAutomated milestone notifications replace manual outreach.")

R += 1
phase_headers(ws2, R)

R += 1  # Support calls
ws2.cell(row=R, column=1, value="\"Where's My Case?\" Calls/Week").font = label_font
add_tooltip(ws2, R, 1, "Client status inquiry calls are pure overhead.\nAutomated status portals and email updates reduce these by 80%.\nRemaining 20% are calls that genuinely need human attention.")
ws2.cell(row=R, column=2).value = f"={RAW['SupportCalls']}"
ws2.cell(row=R, column=3).value = f"={RAW['SupportCalls']}*0.2"
ws2.cell(row=R, column=3).number_format = '0.0'
p3_calls = R

R += 1  # Calls reduced
ws2.cell(row=R, column=1, value="Calls Eliminated/Week").font = label_font
ws2.cell(row=R, column=4).value = f"=B{p3_calls}-C{p3_calls}"
ws2.cell(row=R, column=4).number_format = '0.0'
p3_eliminated = R

R += 1  # Time saved from calls (12 min each)
ws2.cell(row=R, column=1, value="Hours Saved from Calls/Week").font = label_font
add_tooltip(ws2, R, 1, "Each status call takes ~12 minutes (lookup + conversation).\nHours saved = Calls eliminated × 12 minutes ÷ 60.")
ws2.cell(row=R, column=4).value = f"=D{p3_eliminated}*12/60"
ws2.cell(row=R, column=4).number_format = '0.0'
p3_call_hours = R

R += 1  # Revenue from call savings
ws2.cell(row=R, column=1, value="Revenue from Call Time Savings").font = label_font
add_tooltip(ws2, R, 1, "Dollar value of time freed from status calls.\nHours saved × Billing Rate × 52 weeks.")
ws2.cell(row=R, column=5).value = f"=D{p3_call_hours}*{RAW['BillingRate']}*52"
ws2.cell(row=R, column=5).number_format = '$#,##0'
p3_rev_calls = R

R += 1  # Manual update time saved
ws2.cell(row=R, column=1, value="Manual Update Time Saved/Attorney/Week").font = label_font
add_tooltip(ws2, R, 1, "Hours per attorney per week spent manually updating clients.\nCurrent: ~2 hours (emails, calls, letters).\nImproved: ~0.4 hours (only edge cases need manual touch).")
ws2.cell(row=R, column=2, value=2).font = data_font
ws2.cell(row=R, column=3, value=0.4).font = data_font
ws2.cell(row=R, column=4).value = f"=(B{R}-C{R})*{RAW['NumAttorneys']}"
ws2.cell(row=R, column=4).number_format = '0.0'
p3_update = R

R += 1  # Revenue from update savings
ws2.cell(row=R, column=1, value="Revenue from Update Time Savings").font = label_font
ws2.cell(row=R, column=5).value = f"=D{p3_update}*{RAW['BillingRate']}*52"
ws2.cell(row=R, column=5).number_format = '$#,##0'
p3_rev_update = R

R += 1  # Phase 3 totals
ws2.cell(row=R, column=1, value="PHASE 3 ANNUAL ROI").font = Font(name="Inter", size=11, bold=True)
ws2.cell(row=R, column=5).value = f"=E{p3_rev_calls}+E{p3_rev_update}"
ws2.cell(row=R, column=5).number_format = '$#,##0'
ws2.cell(row=R, column=5).fill = summary_fill
ws2.cell(row=R, column=5).font = Font(name="Inter", size=11, bold=True)
p3_annual = R

R += 1
ws2.cell(row=R, column=1, value="Assumptions: 80% support call reduction, 12 min/call, 2→0.4 hrs/attorney/week manual updates.").font = note_font
ws2.merge_cells(f"A{R}:E{R}")

# --- PHASE 4: Marketing Engine ---
R += 2
ws2.cell(row=R, column=1, value="PHASE 4: Marketing Engine (Month 4+)")
style_section(ws2, R, 5)
ws2.merge_cells(f"A{R}:E{R}")
add_tooltip(ws2, R, 1, "Phase 4 adds marketing automation and attribution tracking.\n40% CPA reduction = more clients from same budget.\nOR reduce spend while maintaining client volume.")

R += 1
phase_headers(ws2, R)

R += 1  # Current CPA (FIXED: uses /100 for conversion rate)
ws2.cell(row=R, column=1, value="Cost Per Acquisition (CPA)").font = label_font
add_tooltip(ws2, R, 1, "How much you spend to acquire each client.\nFormula: Marketing Spend ÷ (Leads × Conversion Rate).\nImproved CPA = 60% of current (40% reduction from better attribution and automation).")
ws2.cell(row=R, column=2).value = f"=IF({PCT_CONV}>0,{RAW['MktSpend']}/({RAW['Leads']}*{PCT_CONV}),0)"
ws2.cell(row=R, column=2).number_format = '$#,##0'
ws2.cell(row=R, column=3).value = f"=B{R}*0.6"
ws2.cell(row=R, column=3).number_format = '$#,##0'
p4_cpa = R

R += 1  # Clients at current
ws2.cell(row=R, column=1, value="Clients/Month at Current CPA").font = label_font
add_tooltip(ws2, R, 1, "Current monthly client acquisitions.\nLeads × Conversion Rate.")
ws2.cell(row=R, column=2).value = f"={RAW['Leads']}*{PCT_CONV}"
ws2.cell(row=R, column=2).number_format = '0.0'
p4_cur_clients = R

R += 1  # Clients at improved
ws2.cell(row=R, column=1, value="Clients/Month at Improved CPA").font = label_font
add_tooltip(ws2, R, 1, "How many clients the same budget buys at the improved CPA.\nMarketing Spend ÷ Improved CPA.")
ws2.cell(row=R, column=3).value = f"=IF(C{p4_cpa}>0,{RAW['MktSpend']}/C{p4_cpa},0)"
ws2.cell(row=R, column=3).number_format = '0.0'
p4_new_clients = R

R += 1  # Additional clients
ws2.cell(row=R, column=1, value="Additional Clients/Month").font = label_font
ws2.cell(row=R, column=4).value = f"=C{p4_new_clients}-B{p4_cur_clients}"
ws2.cell(row=R, column=4).number_format = '0.0'
p4_add = R

R += 1  # Revenue
ws2.cell(row=R, column=1, value="Revenue from Additional Clients").font = label_font
add_tooltip(ws2, R, 1, "Dollar value of additional clients.\nAdditional Clients × Average Revenue Per Case.\nAnnual = Monthly × 12.")
ws2.cell(row=R, column=4).value = f"=D{p4_add}*{RAW['AvgRevPerCase']}"
ws2.cell(row=R, column=4).number_format = '$#,##0'
ws2.cell(row=R, column=5).value = f"=D{R}*12"
ws2.cell(row=R, column=5).number_format = '$#,##0'
p4_rev = R

R += 1  # Phase 4 totals
ws2.cell(row=R, column=1, value="PHASE 4 ANNUAL ROI").font = Font(name="Inter", size=11, bold=True)
ws2.cell(row=R, column=5).value = f"=E{p4_rev}"
ws2.cell(row=R, column=5).number_format = '$#,##0'
ws2.cell(row=R, column=5).fill = summary_fill
ws2.cell(row=R, column=5).font = Font(name="Inter", size=11, bold=True)
p4_annual = R

R += 1
ws2.cell(row=R, column=1, value="Assumptions: 40% CPA reduction from attribution tracking and automated follow-up. Marketing spend held constant.").font = note_font
ws2.merge_cells(f"A{R}:E{R}")

# Disclaimer
R += 2
ws2.cell(row=R, column=1, value="Disclaimer: All projections are estimates. Actual results vary based on implementation quality and market conditions. Not financial or legal advice.").font = note_font
ws2.merge_cells(f"A{R}:E{R}")

# ============================================================
# TAB 3: Summary Dashboard
# ============================================================
ws3 = wb.create_sheet("Summary Dashboard")
ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 22
ws3.column_dimensions["E"].width = 16

ws3["A1"] = "Summary Dashboard"
style_header(ws3, 1, 5)
ws3.merge_cells("A1:E1")

# ROI Summary
ws3.cell(row=3, column=1, value="Phase").font = Font(name="Inter", size=10, bold=True, color=LIGHT_TEXT)
ws3.cell(row=3, column=2, value="Annual ROI").font = Font(name="Inter", size=10, bold=True, color=LIGHT_TEXT)
for c in range(1, 3):
    ws3.cell(row=3, column=c).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

phases = [
    ("Phase 1: Intake + Deadlines", p1_annual),
    ("Phase 2: Document + Billing", p2_annual),
    ("Phase 3: Client Communication", p3_annual),
    ("Phase 4: Marketing Engine", p4_annual),
]
for i, (name, annual_row) in enumerate(phases):
    rr = 4 + i
    ws3.cell(row=rr, column=1, value=name).font = label_font
    ws3.cell(row=rr, column=2).value = f"='Phase-by-Phase ROI'!E{annual_row}"
    ws3.cell(row=rr, column=2).number_format = '$#,##0'

# Total
ws3.cell(row=9, column=1, value="TOTAL ANNUAL ROI").font = Font(name="Inter", size=12, bold=True)
ws3.cell(row=9, column=2).value = "=SUM(B4:B7)"
ws3.cell(row=9, column=2).number_format = '$#,##0'
ws3.cell(row=9, column=2).fill = summary_fill
ws3.cell(row=9, column=2).font = Font(name="Inter", size=12, bold=True)
add_tooltip(ws3, 9, 2, "Sum of all 4 phase annual ROI projections.\nThis is the total estimated annual benefit from full automation implementation.")

ws3.cell(row=10, column=1, value="TOTAL MONTHLY ROI").font = Font(name="Inter", size=11, bold=True)
ws3.cell(row=10, column=2).value = "=B9/12"
ws3.cell(row=10, column=2).number_format = '$#,##0'

# Implementation Investment
ws3.cell(row=12, column=1, value="IMPLEMENTATION INVESTMENT")
style_section(ws3, 12, 5)
ws3.merge_cells("A12:E12")

headers = ["Scenario", "Cost", "Annual ROI", "Payback (Months)", "ROI Multiple"]
tips = [
    "Implementation approach",
    "One-time implementation cost",
    "Projected annual return (from above)",
    "Months until ROI exceeds cost. Lower = faster payback.",
    "Annual ROI ÷ Cost. Higher = better return on investment.\nExample: 10x means $10 returned for every $1 invested.",
]
for i, (h, tip) in enumerate(zip(headers, tips), 1):
    cell = ws3.cell(row=13, column=i, value=h)
    cell.font = Font(name="Inter", size=10, bold=True, color=LIGHT_TEXT)
    cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    cell.comment = Comment(tip, AUTHOR, width=250, height=80)

scenarios = [
    ("Low-End (DIY + Tools)", 5000, "You do it yourself using off-the-shelf tools (Zapier, Clio, etc.)"),
    ("Mid-Range (Partial Custom)", 10000, "Mix of templates and custom work. Some consulting help."),
    ("High-End (Full Custom Build)", 15000, "Fully custom automation system built by a specialist."),
]
for i, (name, cost, tip) in enumerate(scenarios):
    rr = 14 + i
    ws3.cell(row=rr, column=1, value=name).font = label_font
    add_tooltip(ws3, rr, 1, tip)
    ws3.cell(row=rr, column=2, value=cost).number_format = '$#,##0'
    ws3.cell(row=rr, column=3).value = "=B9"
    ws3.cell(row=rr, column=3).number_format = '$#,##0'
    ws3.cell(row=rr, column=4).value = f"=IF(B9>0,B{rr}/(B9/12),\"N/A\")"
    ws3.cell(row=rr, column=4).number_format = '0.0'
    ws3.cell(row=rr, column=5).value = f"=IF(B{rr}>0,B9/B{rr},\"N/A\")"
    ws3.cell(row=rr, column=5).number_format = '0.0"x"'

# Practice-Area Benchmarks — ALL WHOLE NUMBERS for percentages
ws3.cell(row=19, column=1, value="PRACTICE-AREA BENCHMARKS (Copy to Your Firm Profile)")
style_section(ws3, 19, 5)
ws3.merge_cells("A19:E19")
add_tooltip(ws3, 19, 1, "These are typical values for each practice area.\nUse them as a starting point if you're unsure about your own numbers.\nCopy a column's values to Tab 1 to see projected ROI for a typical firm.")

benchmarks = {
    "Bankruptcy": [300, 2, 2500, 70, 35, 7, 25, 36, 13, 1500, 18, 8, 2],
    "Criminal Defense": [400, 1, 5000, 65, 25, 8, 18, 48, 5, 2500, 15, 6, 1],
    "Administrative": [350, 3, 4000, 75, 30, 6, 15, 24, 8, 1800, 12, 5, 1],
}

field_names = [
    "Billing Rate ($/hr)", "Attorneys", "Avg Revenue/Case", "Billable Capture %",
    "Caseload", "Admin Hrs Lost/Atty/Wk", "Leads/Month", "Response Time (hrs)",
    "Conversion Rate %", "Marketing Spend/Mo", "Support Calls/Wk", "Billing Hrs/Wk", "Missed Deadlines/Yr",
]

col_start = 20
ws3.cell(row=col_start, column=1, value="Metric").font = Font(name="Inter", size=10, bold=True)
for j, practice in enumerate(benchmarks.keys()):
    ws3.cell(row=col_start, column=2 + j, value=practice).font = Font(name="Inter", size=10, bold=True)

for i, field in enumerate(field_names):
    rr = col_start + 1 + i
    ws3.cell(row=rr, column=1, value=field).font = data_font
    for j, (practice, vals) in enumerate(benchmarks.items()):
        cell = ws3.cell(row=rr, column=2 + j, value=vals[i])
        cell.font = data_font
        if "$/hr" in field or "Spend" in field or "Revenue" in field:
            cell.number_format = '$#,##0'

# Note about percentages
pct_note_row = col_start + 1 + len(field_names) + 1
ws3.cell(row=pct_note_row, column=1, value="Note: Percentages are whole numbers (70 = 70%, 22 = 22%). Enter them exactly as shown into Your Firm Profile.").font = Font(name="Inter", size=9, bold=True, italic=True, color="CC0000")
ws3.merge_cells(f"A{pct_note_row}:E{pct_note_row}")

# Disclaimer
dr = pct_note_row + 2
ws3.cell(row=dr, column=1, value="Disclaimer: All projections are estimates. Not financial, legal, or business advice. The Innovative Native LLC makes no warranties regarding accuracy. Consult qualified professionals.").font = note_font
ws3.merge_cells(f"A{dr}:E{dr}")

ws3.cell(row=dr + 1, column=1, value="© 2026 The Innovative Native LLC").font = note_font

# Save
output = "/Users/makwa/theinnovativenative/projects/002-stan-store-lawfirm-funnel/assets/roi-calculator-automated-lawfirm-blueprint.xlsx"
wb.save(output)
print(f"Created: {output}")
print("v2: All percentages accept whole numbers. Hover tooltips on every field.")
